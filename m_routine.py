import json
import os
import re
import sys

import PySide2
import pandas as pd
import pymysql
from PySide2.QtCore import QDateTime, Qt, QStandardPaths
from PySide2.QtGui import QIcon
from PySide2.QtUiTools import QUiLoader
from PySide2.QtWidgets import QApplication, QFileDialog, QTableWidgetItem
from openpyxl import load_workbook, Workbook

from setting import Setting

dirname = os.path.dirname(PySide2.__file__)
plugin_path = os.path.join(dirname, 'plugins', 'platforms')
os.environ['QT_QPA_PLATFORM_PLUGIN_PATH'] = plugin_path


class BasCheck:
	def __init__(self):
		self.ui = QUiLoader().load('BAS_check.ui')
		self.setting = Setting()
		self.initUI()
		self.ui.dateTimeEdit_1.setDateTime(QDateTime.currentDateTime().addDays(-1))
		self.ui.dateTimeEdit_2.setDateTime(QDateTime.currentDateTime())
		self.ui.pushButton_setting.clicked.connect(self.setting.show)
		self.ui.pushButton_start.clicked.connect(self.startmatch)
		self.ui.pushButton.clicked.connect(self.selectfile)
		self.ui.checkBox.stateChanged.connect(self.usedb)
		self.ui.pushButton_6.clicked.connect(self.export)

	def initUI(self):
		# 图标
		self.ui.pushButton_setting.setIcon(QIcon('设置.png'))
		# 初始化车站
		with open('configure.json', encoding='utf-8', mode='r') as f:
			json_data = json.load(f)
			self.ui.comboBox.addItems(json_data['station'])

	def selectfile(self):
		filename, _ = QFileDialog.getOpenFileName(caption="选择文件",
		                                          dir=QStandardPaths.writableLocation(QStandardPaths.DesktopLocation),
		                                          filter="记录 文件(*.csv)")
		self.ui.lineEdit.setText(filename)

	def export(self):
		export_dir = QFileDialog.getExistingDirectory(caption="选择保存位置",
		                                              dir=QStandardPaths.writableLocation(
			                                              QStandardPaths.DesktopLocation))
		print(export_dir)
		wb = Workbook()
		sheet = wb.active
		sheet.title = '无记录设备'
		sheet.cell(1, 1).value = '设备编号'
		sheet.cell(1, 2).value = '设备描述'
		for i in range(0, self.ui.tableWidget.rowCount()):
			sheet.cell(2 + i, 1).value = self.ui.tableWidget.item(i, 0).text()
			sheet.cell(2 + i, 2).value = self.ui.tableWidget.item(i, 1).text()
		wb.save(export_dir + "\无记录设备.xlsx")

	def usedb(self):
		if self.ui.checkBox.checkState() == Qt.Checked:
			self.ui.lineEdit.setEnabled(False)
			self.ui.pushButton.setEnabled(False)
		else:
			self.ui.lineEdit.setEnabled(True)
			self.ui.pushButton.setEnabled(True)

	def startmatch(self):
		if self.ui.checkBox.checkState() == Qt.Checked:
			# 获取匹配队列
			pattern1 = self.setting.getchecked()  # 类
			pattern2 = {}  # 编号及描述
			pattern2_temp_keys = []
			wb_p = load_workbook(self.ui.comboBox.currentText() + '.xlsx')
			ws_p = wb_p.active
			for i in range(1, ws_p.max_row):
				if ws_p.cell(i, 1).value in pattern1 and ws_p.cell(i, 1).fill.fgColor.rgb != 'FFFF0000':
					# print(ws_p.cell(i, 1).fill.fgColor.rgb)
					pattern2[ws_p.cell(i, 3).value] = ws_p.cell(i, 6).value + ws_p.cell(i, 4).value
			print(pattern2)
			# 使用数据库
			date1 = self.ui.dateTimeEdit_1.date().toString('yyyyMMdd')
			date2 = self.ui.dateTimeEdit_2.date().toString('yyyyMMdd')
			time1 = self.ui.dateTimeEdit_1.time().toString('hhmmss') + '000'
			time2 = self.ui.dateTimeEdit_2.time().toString('hhmmss') + '999'
			mysql_con = pymysql.connect(host='127.0.0.1', user='root', password='iscs200', port=3306, db='xopenshdb')
			sql = 'SELECT * FROM his_event_tab WHERE YYMMDD>=' + date1 + ' AND YYMMDD<=' + date2 + ' AND HHMMSSMS>=' + time1 + ' AND HHMMSSMS<=' + time2
			df_data = pd.read_sql(sql, mysql_con)
			for i in df_data['Event_Desc']:
				for j in pattern2.keys():
					if j in i:
						pattern2_temp_keys.append(j)
			for p in set(pattern2_temp_keys):
				pattern2.pop(p)
			self.ui.tableWidget.clearContents()
			for i in pattern2.keys():
				row = self.ui.tableWidget.rowCount()
				self.ui.tableWidget.setRowCount(row + 1)
				self.ui.tableWidget.setItem(row, 0, QTableWidgetItem(i))
				self.ui.tableWidget.setItem(row, 1, QTableWidgetItem(pattern2[i]))
		else:
			# 本地文件匹配
			datetime1 = QDateTime.fromString(
				self.ui.dateTimeEdit_1.date().toString('yyyy-MM-dd') + self.ui.dateTimeEdit_1.time().toString(
					' hh:mm:ss'), 'yyyy-MM-dd hh:mm:ss')
			datetime2 = QDateTime.fromString(
				self.ui.dateTimeEdit_2.date().toString('yyyy-MM-dd') + self.ui.dateTimeEdit_2.time().toString(
					' hh:mm:ss'), 'yyyy-MM-dd hh:mm:ss')
			# 使用本地文件
			if self.ui.lineEdit.text().endswith('.csv'):
				# 索引从0开始
				# 获取匹配队列
				pattern1 = self.setting.getchecked()  # 类
				pattern2 = {}  # 编号及描述
				pattern2_temp_keys = []
				wb_p = load_workbook(self.ui.comboBox.currentText() + '.xlsx')
				ws_p = wb_p.active
				for i in range(1, ws_p.max_row):
					if ws_p.cell(i, 1).value in pattern1 and ws_p.cell(i, 1).fill.fgColor.rgb != 'FFFF0000':
						print(ws_p.cell(i, 1).fill.fgColor.rgb)
						pattern2[ws_p.cell(i, 3).value] = ws_p.cell(i, 6).value + ws_p.cell(i, 4).value
				# 开始匹配 处理CSV
				# TODO 需细化切换箱和照明
				data = pd.read_csv(self.ui.lineEdit.text(), header=None, encoding='gb2312', names=range(8))
				for i in range(1, data.shape[0]):
					# 时间转换
					datetime = QDateTime.fromString(
						re.search(r'\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}', data[1][i]).group(), 'yyyy-MM-dd hh:mm:ss')
					if datetime1 <= datetime and datetime2 >= datetime:
						for j in pattern2.keys():
							if re.search(j, data[5][i]):
								pattern2_temp_keys.append(j)
				# print(pattern2_temp_keys)
				for key in set(pattern2_temp_keys):
					pattern2.pop(key)
				# print(pattern2)
				self.ui.tableWidget.clearContents()
				for i in pattern2.keys():
					row = self.ui.tableWidget.rowCount()
					self.ui.tableWidget.setRowCount(row + 1)
					self.ui.tableWidget.setItem(row, 0, QTableWidgetItem(i))
					self.ui.tableWidget.setItem(row, 1, QTableWidgetItem(pattern2[i]))

	def show(self):
		self.ui.show()


if __name__ == "__main__":
	app = QApplication([])
	window = BasCheck()
	window.show()
	sys.exit(app.exec_())
